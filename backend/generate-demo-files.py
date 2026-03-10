#!/usr/bin/env python3
"""
Generate all DACP Construction demo files.
Excel files via openpyxl, PDF files via fpdf2.
"""

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from fpdf import FPDF

BASE_DIR = "/Users/teoblind/final/backend/demo-files"
PRICING_JSON = "/Users/teoblind/final/backend/src/data/dacp/pricing_master.json"

# Colors
DARK_BLUE = "1E3A5F"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MEDIUM_GRAY = "D9D9D9"

# Common styles
header_font_large = Font(name="Calibri", size=14, bold=True)
header_font_medium = Font(name="Calibri", size=12, bold=True)
header_font_small = Font(name="Calibri", size=10, bold=True)
body_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
col_header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
col_header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
bottom_border = Border(bottom=Side(style="medium"))
currency_fmt = '$#,##0.00'
number_fmt = '#,##0'


def ensure_dirs():
    for sub in ["estimates", "pricing", "meetings", "reports", "field"]:
        os.makedirs(os.path.join(BASE_DIR, sub), exist_ok=True)


def setup_estimate_workbook(ws, project_name, gc_name, contact, bid_due, estimate_no,
                             line_items, total_bid, confidence_text,
                             exclusions=None, notes=None, extra_header_info=None):
    """Set up a standard estimate worksheet."""
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Row 1-3: Company header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DACP Construction LLC"
    ws['A1'].font = header_font_large
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:E2')
    ws['A2'] = "4200 Westheimer Rd, Suite 200 | Houston, TX 77027"
    ws['A2'].font = Font(name="Calibri", size=10)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:E3')
    ws['A3'] = "Phone: (713) 555-0142 | Email: estimating@dacpconstruction.com"
    ws['A3'].font = Font(name="Calibri", size=10)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Row 4: blank

    # Row 5-6: Project info
    ws['A5'] = f"Project: {project_name}"
    ws['A5'].font = header_font_medium
    ws.merge_cells('A5:C5')

    ws['D5'] = f"Estimate #: {estimate_no}"
    ws['D5'].font = bold_font
    ws.merge_cells('D5:E5')

    ws['A6'] = f"GC: {gc_name} | Contact: {contact}"
    ws['A6'].font = body_font
    ws.merge_cells('A6:C6')

    ws['D6'] = f"Bid Due: {bid_due}"
    ws['D6'].font = bold_font
    ws.merge_cells('D6:E6')

    if extra_header_info:
        ws.merge_cells('A7:E7')
        ws['A7'] = extra_header_info
        ws['A7'].font = Font(name="Calibri", size=10, italic=True, bold=True)
        header_row = 9
    else:
        header_row = 8

    # Column headers
    headers = ["Item Description", "Qty", "Unit", "Unit Price", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center") if col_idx > 1 else Alignment(horizontal="left")
        cell.border = thin_border

    # Data rows
    data_start = header_row + 1
    for i, item in enumerate(line_items):
        row = data_start + i
        fill = light_gray_fill if i % 2 == 0 else white_fill

        ws.cell(row=row, column=1, value=item[0]).font = body_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border

        qty_cell = ws.cell(row=row, column=2, value=item[1])
        qty_cell.font = body_font
        qty_cell.fill = fill
        qty_cell.border = thin_border
        qty_cell.alignment = Alignment(horizontal="center")
        if isinstance(item[1], (int, float)):
            qty_cell.number_format = number_fmt

        ws.cell(row=row, column=3, value=item[2]).font = body_font
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        price_cell = ws.cell(row=row, column=4, value=item[3])
        price_cell.font = body_font
        price_cell.fill = fill
        price_cell.border = thin_border
        price_cell.number_format = currency_fmt
        price_cell.alignment = Alignment(horizontal="right")

        total_cell = ws.cell(row=row, column=5, value=item[4])
        total_cell.font = body_font
        total_cell.fill = fill
        total_cell.border = thin_border
        if isinstance(item[4], (int, float)):
            total_cell.number_format = currency_fmt
        total_cell.alignment = Alignment(horizontal="right")

    # Subtotal / Total rows
    total_row = data_start + len(line_items) + 1

    # Calculate subtotal
    subtotal = sum(item[4] for item in line_items if isinstance(item[4], (int, float)))

    if subtotal != total_bid:
        ws.cell(row=total_row, column=4, value="Subtotal:").font = bold_font
        ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
        sub_cell = ws.cell(row=total_row, column=5, value=subtotal)
        sub_cell.font = bold_font
        sub_cell.number_format = currency_fmt
        sub_cell.alignment = Alignment(horizontal="right")
        sub_cell.border = thin_border
        total_row += 1

    ws.cell(row=total_row, column=4, value="TOTAL BID:").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    bid_cell = ws.cell(row=total_row, column=5, value=total_bid)
    bid_cell.font = Font(name="Calibri", size=11, bold=True)
    bid_cell.number_format = currency_fmt
    bid_cell.alignment = Alignment(horizontal="right")
    bid_cell.border = Border(top=Side(style="double"), bottom=Side(style="double"))

    current_row = total_row + 1

    # Confidence
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1, value=confidence_text).font = Font(name="Calibri", size=10, italic=True)

    # Exclusions
    if exclusions:
        current_row += 2
        ws.cell(row=current_row, column=1, value="EXCLUSIONS:").font = bold_font
        for excl in exclusions:
            current_row += 1
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.cell(row=current_row, column=1, value=f"  \u2022 {excl}").font = body_font

    # Notes
    if notes:
        current_row += 2
        ws.cell(row=current_row, column=1, value="NOTES:").font = bold_font
        for note in notes:
            current_row += 1
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.cell(row=current_row, column=1, value=f"  \u2022 {note}").font = body_font

    # Footer
    current_row += 2
    ws.merge_cells(f'A{current_row}:E{current_row}')
    footer_cell = ws.cell(row=current_row, column=1, value="Prepared by DACP Construction \u2014 Powered by Coppice")
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    footer_cell.alignment = Alignment(horizontal="center")

    return ws


def create_estimate_1():
    """Bishop Arts Mixed-Use"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    line_items = [
        ('Slab-on-grade 6" (4000 PSI)', 45000, 'SF', 14.26, 641700.00),
        ('Curb & gutter (standard 6"x18")', 2800, 'LF', 26.45, 74060.00),
        ('#4 rebar (supply + install)', 8200, 'LF', 1.85, 15170.00),
        ('Wire mesh 6x6 W2.9xW2.9', 45000, 'SF', 0.85, 38250.00),
        ('Concrete finishing (broom)', 45000, 'SF', 1.20, 54000.00),
        ('Mobilization', 1, 'LS', 3500.00, 3500.00),
        ('Concrete testing (3rd party)', 6, 'EA', 450.00, 2700.00),
    ]

    setup_estimate_workbook(
        ws,
        project_name="Bishop Arts Mixed-Use",
        gc_name="Rogers-O'Brien",
        contact="David Kim (dkim@rogers-obrien.com)",
        bid_due="March 19, 2026",
        estimate_no="EST-2026-001",
        line_items=line_items,
        total_bid=847300.00,
        confidence_text="Confidence: 92% | Based on 3 comparable projects",
        exclusions=[
            "Subgrade preparation and compaction by others",
            "Concrete material (furnished by GC or billed separately at TXI rates)",
            "Dewatering if required",
            "Permits and inspections",
        ],
        notes=[
            "Pricing valid for 30 days from date of estimate",
            "Based on TXI March 2026 pricing ($158/CY for 4000 PSI)",
        ],
    )

    path = os.path.join(BASE_DIR, "estimates", "DACP_Estimate_BishopArts_MixedUse.xlsx")
    wb.save(path)
    return path


def create_estimate_2():
    """I-35 Retaining Walls"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    line_items = [
        ('Retaining wall (cantilever)', 2150, 'LF', 145.00, 311750.00),
        ('Grade beam 18"x24"', 400, 'LF', 52.00, 20800.00),
        ('#5 rebar (supply + install)', 6400, 'LF', 2.40, 15360.00),
        ('Form & strip (wall, two-sided)', 8600, 'SFCA', 18.00, 154800.00),
        ('Concrete testing (3rd party)', 8, 'EA', 450.00, 3600.00),
        ('Mobilization', 1, 'LS', 5000.00, 5000.00),
    ]

    setup_estimate_workbook(
        ws,
        project_name="I-35 Retaining Walls",
        gc_name="Hensel Phelps",
        contact="Lisa Chen (lchen@henselphelps.com)",
        bid_due="March 21, 2026",
        estimate_no="EST-2026-002",
        line_items=line_items,
        total_bid=312000.00,
        confidence_text="Confidence: 78% \u2014 First bid with Hensel Phelps",
    )

    path = os.path.join(BASE_DIR, "estimates", "DACP_Estimate_I35_RetainingWalls.xlsx")
    wb.save(path)
    return path


def create_estimate_3():
    """Memorial Hermann Phase 2"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    line_items = [
        ('Slab-on-grade 6" (4000 PSI)', 8500, 'SF', 14.26, 121210.00),
        ('Curb & gutter (standard)', 1200, 'LF', 26.45, 31740.00),
        ('Sidewalk 6" (ADA compliant)', 2250, 'SF', 11.00, 24750.00),
        ('#4 rebar (supply + install)', 4800, 'LF', 1.85, 8880.00),
        ('Wire mesh 6x6', 8500, 'SF', 0.85, 7225.00),
        ('Mobilization', 1, 'LS', 3500.00, 3500.00),
        ('Concrete testing (3rd party)', 4, 'EA', 450.00, 1800.00),
    ]

    setup_estimate_workbook(
        ws,
        project_name="Memorial Hermann Phase 2",
        gc_name="Turner Construction",
        contact="Mike Rodriguez (mrodriguez@turner.com)",
        bid_due="March 14, 2026",
        estimate_no="EST-2026-003",
        line_items=line_items,
        total_bid=266000.00,
        confidence_text="Confidence: 92% \u2014 Based on Phase 1 actuals",
    )

    path = os.path.join(BASE_DIR, "estimates", "DACP_Estimate_MemorialHermann_Ph2.xlsx")
    wb.save(path)
    return path


def create_estimate_4():
    """Samsung Fab Expansion — Equipment Pads (REVISED)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    line_items = [
        ('Slab-on-grade 8" (4000 PSI) equip pads', 60, 'EA', 1850.00, 111000.00),
        ('Grade beam 18"x24"', 800, 'LF', 52.00, 41600.00),
        ('Slab-on-grade 6" (4000 PSI)', 6000, 'SF', 14.26, 85560.00),
        ('Mobilization', 1, 'LS', 3500.00, 3500.00),
        ('Concrete testing (3rd party)', 4, 'EA', 450.00, 1800.00),
    ]

    setup_estimate_workbook(
        ws,
        project_name="Samsung Fab Expansion \u2014 Equipment Pads (REVISED)",
        gc_name="DPR Construction",
        contact="Sarah Williams (swilliams@dpr.com)",
        bid_due="March 20, 2026",
        estimate_no="EST-2026-004-R1",
        line_items=line_items,
        total_bid=185000.00,
        confidence_text="Original: $165,000 \u2192 Revised: $185,000 (+12%)",
        notes=[
            "Revised from 45 to 60 equipment pads per DPR scope change",
            "TXI March 2026 rates applied ($158/CY for 4000 PSI)",
        ],
        extra_header_info="REVISED ESTIMATE \u2014 Original: $165,000 \u2192 Revised: $185,000 (+12%)",
    )

    path = os.path.join(BASE_DIR, "estimates", "DACP_Estimate_SamsungFab_Revised.xlsx")
    wb.save(path)
    return path


def create_estimate_5():
    """McKinney Town Center (DRAFT)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Row 1-3: Company header
    ws.merge_cells('A1:E1')
    ws['A1'] = "DACP Construction LLC"
    ws['A1'].font = header_font_large
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:E2')
    ws['A2'] = "4200 Westheimer Rd, Suite 200 | Houston, TX 77027"
    ws['A2'].font = Font(name="Calibri", size=10)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:E3')
    ws['A3'] = "Phone: (713) 555-0142 | Email: estimating@dacpconstruction.com"
    ws['A3'].font = Font(name="Calibri", size=10)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Row 5-6: Project info
    ws['A5'] = "Project: McKinney Town Center"
    ws['A5'].font = header_font_medium
    ws.merge_cells('A5:C5')
    ws['D5'] = "Estimate #: EST-2026-005"
    ws['D5'].font = bold_font
    ws.merge_cells('D5:E5')

    ws['A6'] = "GC: Austin Commercial | Contact: estimating@austincommercial.com"
    ws['A6'].font = body_font
    ws.merge_cells('A6:C6')
    ws['D6'] = "Bid Due: March 25, 2026"
    ws['D6'].font = bold_font
    ws.merge_cells('D6:E6')

    # DRAFT warning
    ws.merge_cells('A7:E7')
    ws['A7'] = "\u26a0\ufe0f  DRAFT \u2014 DO NOT SEND  \u26a0\ufe0f"
    ws['A7'].font = Font(name="Calibri", size=14, bold=True, color="FF0000")
    ws['A7'].alignment = Alignment(horizontal="center")
    ws['A7'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Column headers row 9
    headers = ["Item Description", "Qty", "Unit", "Unit Price", "Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center") if col_idx > 1 else Alignment(horizontal="left")
        cell.border = thin_border

    # Draft line items
    draft_items = [
        ('Slab-on-grade 6" (qty TBD)', '???', 'SF', 14.26, 'TBD'),
        ('Curb & gutter (qty TBD)', '???', 'LF', 26.45, 'TBD'),
        ('Drilled piers 18" dia (qty TBD)', '???', 'EA', 1200.00, 'TBD'),
    ]

    for i, item in enumerate(draft_items):
        row = 10 + i
        fill = light_gray_fill if i % 2 == 0 else white_fill

        ws.cell(row=row, column=1, value=item[0]).font = body_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=item[1]).font = body_font
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=3, value=item[2]).font = body_font
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        price_cell = ws.cell(row=row, column=4, value=item[3])
        price_cell.font = body_font
        price_cell.fill = fill
        price_cell.border = thin_border
        price_cell.number_format = currency_fmt
        price_cell.alignment = Alignment(horizontal="right")

        ws.cell(row=row, column=5, value=item[4]).font = body_font
        ws.cell(row=row, column=5).fill = fill
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

    # Total row
    ws.cell(row=14, column=4, value="TOTAL BID:").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=14, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=14, column=5, value="TBD").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=14, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=14, column=5).border = Border(top=Side(style="double"), bottom=Side(style="double"))

    # Status note
    ws.merge_cells('A16:E16')
    ws.cell(row=16, column=1, value="Status: 48-page spec document being parsed. 3 of estimated 8+ line items identified.").font = Font(name="Calibri", size=10, italic=True)

    # Footer
    ws.merge_cells('A18:E18')
    footer_cell = ws.cell(row=18, column=1, value="Prepared by DACP Construction \u2014 Powered by Coppice")
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    footer_cell.alignment = Alignment(horizontal="center")

    path = os.path.join(BASE_DIR, "estimates", "DACP_Estimate_McKinneyTC_Draft.xlsx")
    wb.save(path)
    return path


def create_pricing_table():
    """Master Pricing Table with data from pricing_master.json"""
    with open(PRICING_JSON, 'r') as f:
        pricing_data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Pricing"

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40

    # Row 1-3: Company header
    ws.merge_cells('A1:H1')
    ws['A1'] = "DACP Construction LLC"
    ws['A1'].font = header_font_large
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = "4200 Westheimer Rd, Suite 200 | Houston, TX 77027"
    ws['A2'].font = Font(name="Calibri", size=10)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:H3')
    ws['A3'] = "Phone: (713) 555-0142 | Email: estimating@dacpconstruction.com"
    ws['A3'].font = Font(name="Calibri", size=10)
    ws['A3'].alignment = Alignment(horizontal="center")

    # Row 5: Title
    ws.merge_cells('A5:H5')
    ws['A5'] = "MASTER PRICING TABLE \u2014 Effective March 2026"
    ws['A5'].font = Font(name="Calibri", size=12, bold=True)
    ws['A5'].alignment = Alignment(horizontal="center")

    # Row 7: Column headers
    headers = ["Category", "Item", "Unit", "Material", "Labor", "Equipment", "Unit Price", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center") if col_idx != 2 and col_idx != 8 else Alignment(horizontal="left")
        cell.border = thin_border

    # Data rows
    for i, item in enumerate(pricing_data):
        row = 8 + i
        fill = light_gray_fill if i % 2 == 0 else white_fill

        values = [
            item["category"],
            item["item"],
            item["unit"],
            item["material_cost"],
            item["labor_cost"],
            item["equipment_cost"],
            item["unit_price"],
            item["notes"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.border = thin_border

            if col_idx in (4, 5, 6, 7):
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    # Supplier pricing section
    supplier_start = 8 + len(pricing_data) + 2
    ws.merge_cells(f'A{supplier_start}:H{supplier_start}')
    ws.cell(row=supplier_start, column=1, value="SUPPLIER PRICING").font = Font(name="Calibri", size=12, bold=True)

    supplier_header_row = supplier_start + 1
    sup_headers = ["Supplier", "Product", "", "Rate", "", "", "", "Effective"]
    for col_idx, header in enumerate(sup_headers, 1):
        if header:
            cell = ws.cell(row=supplier_header_row, column=col_idx, value=header)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Merge some header cells for supplier section
    ws.merge_cells(f'B{supplier_header_row}:C{supplier_header_row}')
    ws.merge_cells(f'D{supplier_header_row}:G{supplier_header_row}')

    supplier_items = [
        ("TXI", "3000 PSI Ready-Mix", "$149.00/CY", "Mar 2026"),
        ("TXI", "4000 PSI Ready-Mix", "$158.00/CY", "Mar 2026"),
        ("TXI", "5000 PSI Ready-Mix", "$172.00/CY", "Mar 2026"),
        ("CMC Steel", "#4 Rebar", "$1,050/ton", "Feb 2026"),
        ("CMC Steel", "#5 Rebar", "$1,080/ton", "Feb 2026"),
    ]

    for i, (supplier, product, rate, effective) in enumerate(supplier_items):
        row = supplier_header_row + 1 + i
        fill = light_gray_fill if i % 2 == 0 else white_fill

        ws.cell(row=row, column=1, value=supplier).font = body_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border

        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value=product).font = body_font
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border

        ws.merge_cells(f'D{row}:G{row}')
        ws.cell(row=row, column=4, value=rate).font = body_font
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=8, value=effective).font = body_font
        ws.cell(row=row, column=8).fill = fill
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")

    # Markups section
    markups_row = supplier_header_row + 1 + len(supplier_items) + 2
    ws.merge_cells(f'A{markups_row}:H{markups_row}')
    ws.cell(row=markups_row, column=1, value="STANDARD MARKUPS").font = Font(name="Calibri", size=12, bold=True)

    markups = [
        ("Overhead", "15%"),
        ("Profit", "10%"),
        ("Bond", "1.5%"),
    ]
    for i, (name, pct) in enumerate(markups):
        row = markups_row + 1 + i
        ws.cell(row=row, column=1, value=name).font = bold_font
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value=pct).font = body_font
        ws.cell(row=row, column=2).border = thin_border

    # Footer
    footer_row = markups_row + 1 + len(markups) + 2
    ws.merge_cells(f'A{footer_row}:H{footer_row}')
    footer_cell = ws.cell(row=footer_row, column=1, value="Prepared by DACP Construction \u2014 Powered by Coppice")
    footer_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    footer_cell.alignment = Alignment(horizontal="center")

    path = os.path.join(BASE_DIR, "pricing", "DACP_MasterPricingTable_2026.xlsx")
    wb.save(path)
    return path


# ──────────────────────────────────────────────────
# PDF Generation
# ──────────────────────────────────────────────────

class DACPReport(FPDF):
    """Base PDF class with DACP header/footer."""

    def __init__(self, doc_type="REPORT"):
        super().__init__()
        self.doc_type = doc_type

    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 8, "DACP Construction LLC", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 10)
        self.cell(100, 5, self.doc_type)
        self.cell(0, 5, "Powered by Coppice", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(30, 58, 95)
        self.set_line_width(0.5)
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", align="L")
        self.cell(0, 10, f"Generated: March 9, 2026", align="R", new_x="LMARGIN")

    def section_title(self, title):
        self.set_font("Helvetica", "B", 12)
        self.set_fill_color(30, 58, 95)
        self.set_text_color(255, 255, 255)
        self.cell(0, 7, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def body_text(self, text):
        self.set_font("Helvetica", "", 10)
        self.multi_cell(0, 5, text)
        self.ln(2)

    def bullet(self, text):
        self.set_font("Helvetica", "", 10)
        self.set_x(self.l_margin)
        self.multi_cell(0, 5, f"  -  {text}")

    def action_item(self, text, checked=False):
        self.set_font("Helvetica", "", 10)
        marker = "[x]" if checked else "[ ]"
        self.set_x(self.l_margin)
        self.multi_cell(0, 5, f"  {marker}  {text}")

    def key_value(self, key, value):
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "B", 10)
        self.cell(50, 5, key, new_x="END")
        self.set_font("Helvetica", "", 10)
        self.multi_cell(0, 5, value)


def create_meeting_turner():
    pdf = DACPReport("MEETING SUMMARY")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Turner Coordination Call", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 6, "Memorial Hermann Phase 2", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.key_value("Date:", "March 6, 2026")
    pdf.key_value("Duration:", "38 minutes")
    pdf.key_value("Participants:", "Mike Rodriguez (Turner), David Castillo (DACP), Marcel Pineda (DACP)")
    pdf.ln(4)

    pdf.section_title("SUMMARY")
    pdf.body_text(
        "Reviewed concrete pour schedule for Memorial Hermann Phase 2. Turner confirmed they want "
        "pours to begin March 24. DACP confirmed crew availability for 3 consecutive pour days."
    )
    pdf.body_text(
        "Mike mentioned Turner is scoping a potential Phase 3 -- an additional parking structure. "
        "He asked DACP to provide preliminary pricing on 36\" drilled piers to 35' depth for the "
        "parking structure foundations."
    )
    pdf.body_text(
        "Discussion touched on the current concrete pricing environment. Mike acknowledged the 8% "
        "increase from TXI but said Turner is absorbing it on this project since the GMP was locked "
        "before the increase."
    )

    pdf.section_title("ACTION ITEMS")
    pdf.action_item("Confirm rebar delivery date with CMC Steel -- Marcel -- by March 10")
    pdf.action_item("Provide Phase 3 pier pricing estimate -- Estimating Bot -- by March 12")
    pdf.action_item("Send updated pour schedule to Turner PM -- David -- by March 8")
    pdf.ln(2)

    pdf.section_title("KEY DECISIONS")
    pdf.bullet("Pours start March 24 (confirmed)")
    pdf.bullet("DACP will provide Phase 3 pricing as a courtesy estimate")

    path = os.path.join(BASE_DIR, "meetings", "Turner_CoordinationCall_Mar6.pdf")
    pdf.output(path)
    return path


def create_meeting_standup():
    pdf = DACPReport("MEETING SUMMARY")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Weekly Team Standup", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.key_value("Date:", "March 5, 2026")
    pdf.key_value("Duration:", "45 minutes")
    pdf.key_value("Participants:", "David Castillo, Marcel Pineda, Carlos Mendez (Foreman), Juan Reyes (Estimator)")
    pdf.ln(4)

    pdf.section_title("SUMMARY")
    pdf.body_text(
        "Weekly operations review. Active job status: Westpark Retail is on track at 40% complete. "
        "St. Luke's Parking pier drilling starts next week -- geotech shows potential rock at 22', "
        "team discussed contingency plan."
    )
    pdf.body_text(
        "Samsung Fab estimate needs revision -- DPR added 15 pads to scope and TXI raised concrete "
        "prices 8% effective March 1. Juan flagged that Bishop Arts RFQ came in from Rogers-O'Brien "
        "with tight timeline."
    )
    pdf.body_text(
        "Marcel reported TXI is not willing to lock Q2 pricing yet. Team agreed to use current $158/CY "
        "for all active estimates."
    )

    pdf.section_title("ACTION ITEMS")
    pdf.action_item("Review Bishop Arts specs -- Juan (done -- bot generated estimate)", checked=True)
    pdf.action_item("Call TXI about Q2 price lock -- Marcel -- by March 7")
    pdf.action_item("Document rock conditions at Frisco Station -- Carlos -- by March 8")
    pdf.action_item("Follow up with Hensel Phelps on I-35 bid -- David -- ongoing")
    pdf.ln(2)

    pdf.section_title("KEY DECISIONS")
    pdf.bullet("Use $158/CY (TXI March rate) for all active estimates")
    pdf.bullet("Samsung Fab revision to include full 60 pads + updated material costs")
    pdf.bullet("Bishop Arts estimate assigned to bot")

    path = os.path.join(BASE_DIR, "meetings", "WeeklyStandup_Mar5.pdf")
    pdf.output(path)
    return path


def create_meeting_dpr():
    pdf = DACPReport("MEETING SUMMARY")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "DPR Samsung Fab Scope Review", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.key_value("Date:", "March 3, 2026")
    pdf.key_value("Duration:", "25 minutes")
    pdf.key_value("Participants:", "Sarah Williams (DPR), David Castillo (DACP)")
    pdf.ln(4)

    pdf.section_title("SUMMARY")
    pdf.body_text(
        "DPR informed DACP that Samsung is expanding equipment pad scope from 45 to 60 units. "
        "Sarah requested revised estimate by end of week."
    )
    pdf.body_text(
        "David raised the concrete price increase -- TXI raised ready-mix 8% effective March 1. "
        "DPR's position: want DACP to absorb half. David pushed back firmly, citing full pass-through "
        "on material costs is industry standard. Sarah said she'd take it to her PM. David offered to "
        "send TXI price letter as documentation."
    )

    pdf.section_title("ACTION ITEMS")
    pdf.action_item("Submit revised estimate with 60 pads + updated costs -- Estimating Bot -- by March 7")
    pdf.action_item("Send DPR the TXI price letter -- Marcel -- by March 5")
    pdf.ln(2)

    pdf.section_title("KEY DECISIONS")
    pdf.bullet("Scope expanded from 45 to 60 equipment pads")
    pdf.bullet("DACP position: full pass-through on material costs")
    pdf.bullet("TXI price letter to be sent as documentation")

    path = os.path.join(BASE_DIR, "meetings", "DPR_SamsungFab_ScopeReview_Mar3.pdf")
    pdf.output(path)
    return path


def create_daily_ops(day):
    """Create daily ops report for March 5-8."""
    pdf = DACPReport("DAILY OPERATIONS REPORT")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, f"Daily Operations Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Date: March {day}, 2026", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "Auto-generated by Coppice Reporting Engine at 6:00 AM CT", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Daily variations
    day_data = {
        5: {
            "westpark": ("40%", "Forms set for section B pour", "On track"),
            "stlukes": ("18%", "Mobilization & layout", "On track"),
            "samsung": ("Pending", "Estimate revision in progress", "Revised +12%"),
            "bishop": ("Pending", "RFQ received from R-O", "Not started"),
            "agents": [
                "Parsed Bishop Arts RFQ specs (48 pages)",
                "Generated initial estimate framework EST-2026-001",
                "Updated TXI pricing in master table (+8%)",
            ],
            "cy": 0, "estimates": 1, "emails": 3,
        },
        6: {
            "westpark": ("40%", "Poured 48 CY section A (4000 PSI)", "On track"),
            "stlukes": ("20%", "Pier drilling started - P-1, P-2", "On track"),
            "samsung": ("Pending", "Revised estimate submitted", "Revised +12%"),
            "bishop": ("Pending", "Estimate generated (EST-2026-001)", "Not started"),
            "agents": [
                "Generated Bishop Arts estimate: $847K (92% confidence)",
                "Submitted Samsung Fab revised estimate: $185K",
                "Drafted Turner coordination call agenda",
                "Updated project pipeline dashboard",
            ],
            "cy": 48, "estimates": 2, "emails": 5,
        },
        7: {
            "westpark": ("42%", "Poured 52 CY section B, cured section A", "On track"),
            "stlukes": ("22%", "Drilled P-3, P-4; rock flagged on P-5 at 28'", "Monitor"),
            "samsung": ("Pending", "Awaiting DPR cost absorption response", "Revised +12%"),
            "bishop": ("Pending", "Estimate sent to Rogers-O'Brien", "Not started"),
            "agents": [
                "Generated I-35 Retaining Walls estimate: $312K",
                "Flagged rock discrepancy at Frisco Station P-5",
                "Sent Bishop Arts estimate to dkim@rogers-obrien.com",
                "Began parsing McKinney Town Center specs",
            ],
            "cy": 52, "estimates": 1, "emails": 4,
        },
        8: {
            "westpark": ("44%", "Finished section B, formed section C", "On track"),
            "stlukes": ("24%", "Drilled P-6; awaiting geotech update on P-5 area", "Monitor"),
            "samsung": ("Pending", "DPR reviewing cost pass-through", "Revised +12%"),
            "bishop": ("Pending", "Awaiting GC response", "Not started"),
            "agents": [
                "Generated Memorial Hermann Phase 2 estimate: $266K",
                "McKinney Town Center parsing: 3/8+ line items identified",
                "Updated daily field logs for Westpark",
                "Sent Memorial Hermann estimate to mrodriguez@turner.com",
                "Compiled weekly metrics summary",
            ],
            "cy": 45, "estimates": 1, "emails": 6,
        },
    }

    d = day_data[day]

    # Active Jobs table
    pdf.section_title("ACTIVE JOBS")

    pdf.set_font("Helvetica", "B", 9)
    col_widths = [40, 20, 70, 28]
    headers = ["Job", "% Complete", "Today's Activity", "Budget Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 6, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    jobs = [
        ("Westpark Retail", d["westpark"][0], d["westpark"][1], d["westpark"][2]),
        ("St. Luke's Parking", d["stlukes"][0], d["stlukes"][1], d["stlukes"][2]),
        ("Samsung Fab", d["samsung"][0], d["samsung"][1], d["samsung"][2]),
        ("Bishop Arts", d["bishop"][0], d["bishop"][1], d["bishop"][2]),
    ]
    for job in jobs:
        for i, val in enumerate(job):
            pdf.cell(col_widths[i], 6, val, border=1)
        pdf.ln()
    pdf.ln(4)

    # Estimating Pipeline
    pdf.section_title("ESTIMATING PIPELINE")
    pdf.set_font("Helvetica", "B", 9)
    pipe_widths = [36, 30, 24, 22, 20]
    pipe_headers = ["Project", "GC", "Status", "Due Date", "Value"]
    for i, h in enumerate(pipe_headers):
        pdf.cell(pipe_widths[i], 6, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    pipeline = [
        ("Bishop Arts", "Rogers-O'Brien", "Estimated", "Mar 19", "$847K"),
        ("I-35 Retaining", "Hensel Phelps", "Estimated", "Mar 21", "$312K"),
        ("McKinney Town Ctr", "Austin Comm", "In Progress", "Mar 25", "TBD"),
        ("Memorial Hermann Ph2", "Turner", "Sent", "Mar 14", "$266K"),
    ]
    for row in pipeline:
        for i, val in enumerate(row):
            pdf.cell(pipe_widths[i], 6, val, border=1)
        pdf.ln()
    pdf.ln(4)

    # Agent Activity
    pdf.section_title("AGENT ACTIVITY (last 24h)")
    for item in d["agents"]:
        pdf.bullet(item)
    pdf.ln(4)

    # Key Metrics
    pdf.section_title("KEY METRICS")
    pdf.set_font("Helvetica", "", 10)
    metrics = (
        f"Crews Deployed: 2   |   CY Poured: {d['cy']}   |   "
        f"Estimates Generated: {d['estimates']}   |   Emails Sent: {d['emails']}"
    )
    pdf.cell(0, 6, metrics, new_x="LMARGIN", new_y="NEXT")

    path = os.path.join(BASE_DIR, "reports", f"DACP_DailyOps_Mar{day}.pdf")
    pdf.output(path)
    return path


def create_field_westpark():
    pdf = DACPReport("FIELD REPORT")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Daily Field Log", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.key_value("Job:", "Westpark Retail Center (J-009)")
    pdf.key_value("Date:", "March 7, 2026")
    pdf.key_value("Foreman:", "Carlos Mendez")
    pdf.key_value("Weather:", "Clear, 78 deg F")
    pdf.ln(4)

    pdf.section_title("WORK PERFORMED")
    pdf.bullet("Poured 52 CY slab section B (4000 PSI)")
    pdf.bullet("Finished and cured section A from yesterday")
    pdf.bullet("Set forms for section C")
    pdf.ln(2)

    pdf.section_title("MATERIALS")
    pdf.bullet("4000 PSI concrete: 52 CY (TXI delivery)")
    pdf.bullet("#5 rebar: 2,400 LB")
    pdf.bullet("Wire mesh 6x6: 3,200 SF")
    pdf.ln(2)

    pdf.section_title("LABOR")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "Crew: 6 workers  |  Hours: 48  |  Overtime: 4 hrs  |  Cost: $4,800", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.section_title("EQUIPMENT")
    pdf.bullet("42m boom pump")
    pdf.bullet("2x vibrators")
    pdf.bullet("Power trowel")
    pdf.ln(2)

    pdf.section_title("ISSUES")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "None", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.section_title("NOTES")
    pdf.body_text("Good pour day. Section B went smooth. On track for section C pour Wednesday.")
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "________________", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "Foreman Signature", new_x="LMARGIN", new_y="NEXT")

    path = os.path.join(BASE_DIR, "field", "Westpark_DailyLog_Mar7.pdf")
    pdf.output(path)
    return path


def create_field_frisco():
    pdf = DACPReport("FIELD REPORT -- ISSUE FLAG")
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Field Report -- Issue Flag", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.key_value("Job:", "Frisco Station Mixed-Use (J-002)")
    pdf.key_value("Date:", "March 7, 2026")
    pdf.key_value("Foreman:", "Carlos Mendez")
    pdf.key_value("Weather:", "Overcast, 72 deg F")
    pdf.ln(4)

    # Warning box
    pdf.set_fill_color(255, 240, 200)
    pdf.set_draw_color(200, 150, 0)
    pdf.set_line_width(0.8)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "  WARNING: Rock encountered at 28' on pier P-5", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_line_width(0.2)
    pdf.ln(4)

    pdf.section_title("DETAILS")
    pdf.body_text(
        "During drilling of pier P-5, drill rig encountered rock formation at 28 feet. Original geotech "
        "report indicated rock at 35+ feet in this area. This is a 7-foot discrepancy from the geotech "
        "boring logs."
    )

    pdf.section_title("IMPACT")
    pdf.bullet("Estimated cost impact: ~$8,000 (additional rock drilling, extended rig time)")
    pdf.bullet("Schedule impact: 1-2 day delay on P-5 and adjacent piers P-6, P-7")
    pdf.bullet("May need to request updated geotech boring logs for the P-5 through P-8 area")
    pdf.ln(2)

    pdf.section_title("RECOMMENDED ACTION")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "1.  Request updated boring logs from geotech consultant", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "2.  Document conditions for change order to DPR", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "3.  Review adjacent pier locations for similar risk", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "________________", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "Foreman Signature", new_x="LMARGIN", new_y="NEXT")

    path = os.path.join(BASE_DIR, "field", "FriscoStation_RockFlag_Mar7.pdf")
    pdf.output(path)
    return path


def create_txi_price_letter():
    pdf = DACPReport("SUPPLIER CORRESPONDENCE")
    pdf.add_page()

    # Override header for this one - use TXI letterhead style
    pdf.set_y(35)

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "TXI (a U.S. Concrete company)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "1341 W Mockingbird Ln, Suite 700W", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "Dallas, TX 75247", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "March 1, 2026", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 5, "Re: Ready-Mix Concrete Price Adjustment -- Effective March 1, 2026", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5,
        "Dear Valued Customer,\n\n"
        "Due to increased raw material costs (cement +6%, aggregates +4%) and rising "
        "transportation/fuel surcharges, we are implementing the following price adjustments "
        "effective March 1, 2026:"
    )
    pdf.ln(4)

    # Price table
    pdf.set_font("Helvetica", "B", 9)
    tw = [55, 35, 35, 20]
    price_headers = ["Product", "Previous Rate", "New Rate", "Change"]
    for i, h in enumerate(price_headers):
        pdf.cell(tw[i], 6, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    price_rows = [
        ("3000 PSI Ready-Mix", "$142.00/CY", "$149.00/CY", "+4.9%"),
        ("4000 PSI Ready-Mix", "$149.00/CY", "$158.00/CY", "+6.0%"),
        ("5000 PSI Ready-Mix", "$163.00/CY", "$172.00/CY", "+5.5%"),
        ("Fiber Mesh Additive", "$12.00/CY", "$13.00/CY", "+8.3%"),
    ]
    for row in price_rows:
        for i, val in enumerate(row):
            pdf.cell(tw[i], 6, val, border=1)
        pdf.ln()
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5,
        "These prices apply to all new orders placed on or after March 1, 2026. Existing contracts "
        "with locked pricing are honored through their stated expiration date.\n\n"
        "We appreciate your continued partnership. Please contact your TXI sales representative "
        "with any questions."
    )
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "Sincerely,", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 5, "Regional Sales Office", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "TXI -- A U.S. Concrete Company", new_x="LMARGIN", new_y="NEXT")

    path = os.path.join(BASE_DIR, "pricing", "TXI_PriceLetter_Mar2026.pdf")
    pdf.output(path)
    return path


def main():
    ensure_dirs()
    generated = []

    print("Generating DACP Construction demo files...\n")

    # Excel files
    print("--- Excel Files ---")
    generated.append(create_estimate_1())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_estimate_2())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_estimate_3())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_estimate_4())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_estimate_5())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_pricing_table())
    print(f"  [OK] {generated[-1]}")

    # PDF files
    print("\n--- PDF Files ---")
    generated.append(create_meeting_turner())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_meeting_standup())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_meeting_dpr())
    print(f"  [OK] {generated[-1]}")

    for day in [5, 6, 7, 8]:
        generated.append(create_daily_ops(day))
        print(f"  [OK] {generated[-1]}")

    generated.append(create_field_westpark())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_field_frisco())
    print(f"  [OK] {generated[-1]}")

    generated.append(create_txi_price_letter())
    print(f"  [OK] {generated[-1]}")

    # Summary
    print(f"\n{'='*60}")
    print(f"Generated {len(generated)} files:\n")
    for f in generated:
        size = os.path.getsize(f)
        if size > 1024:
            size_str = f"{size/1024:.1f} KB"
        else:
            size_str = f"{size} B"
        print(f"  {f}  ({size_str})")

    print(f"\nDone!")


if __name__ == "__main__":
    main()
